from flask import Flask, request
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

FILE = "gastos.xlsx"
SHEET = "Movimientos"

# Mapeo opcional: cambiá por sus números reales (formato Twilio: whatsapp:+54...)
ALIAS = {
    # "whatsapp:+54911XXXXXXXX": "Facu",
    # "whatsapp:+54911YYYYYYYY": "Novi",
}

def registrar_gasto(monto: int, categoria: str, descripcion: str, who: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(FILE):
        wb = load_workbook(FILE)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb.create_sheet(SHEET)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET
        ws.append(["fecha", "monto", "categoria", "descripcion", "quien", "from_raw"])

    nombre = ALIAS.get(who, who)
    ws.append([ts, monto, categoria.lower(), descripcion, nombre, who])
    wb.save(FILE)

def parsear(texto: str):
    # Formato recomendado: gasto <monto> <categoria> [descripcion]
    parts = texto.strip().split()
    if len(parts) < 3 or parts[0].lower() != "gasto":
        raise ValueError("Usá: gasto <monto> <categoria> [descripcion]. Ej: gasto 10000 supermercado")
    monto = int(parts[1].replace(".", "").replace(",", ""))
    categoria = parts[2]
    descripcion = " ".join(parts[3:]) if len(parts) > 3 else ""
    return monto, categoria, descripcion

@app.route("/twilio", methods=["POST"])
def inbound():
    texto = request.form.get("Body", "")
    who = request.form.get("From", "")  # <- identifica quién mandó el mensaje

    try:
        monto, cat, desc = parsear(texto)
        registrar_gasto(monto, cat, desc, who)

        return f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
  <Message>✅ Registrado: ${monto} en {cat}. {('(' + desc + ')') if desc else ''}</Message>
</Response>""", 200, {"Content-Type": "application/xml"}

    except Exception as e:
        return f"""<?xml version="1.0" encoding="UTF-8"?>
<Response>
  <Message>❌ {str(e)}</Message>
</Response>""", 200, {"Content-Type": "application/xml"}

if __name__ == "__main__":
    app.run(port=5000)
